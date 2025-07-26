"""
Results tab for the Brightness Detector application
"""

import os
import json
from datetime import datetime
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
    QLabel, QTableWidget, QTableWidgetItem, QGroupBox, 
    QTextEdit, QFileDialog, QMessageBox, QSplitter,
    QFrame, QSlider, QGraphicsView, QGraphicsScene, QGraphicsPixmapItem,
    QInputDialog,
)
from PyQt5.QtCore import Qt, pyqtSlot
from PyQt5.QtGui import QImage, QPixmap, QPainter, QPen, QColor
import numpy as np
import cv2
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

# Replace pandas with openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available. Excel export will be disabled.")


class PhotoViewer(QGraphicsView):
    """A QGraphicsView for displaying and interacting with images."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.scene = QGraphicsScene(self)
        self.setScene(self.scene)
        self.photo_item = QGraphicsPixmapItem()
        self.scene.addItem(self.photo_item)
        
        # Placeholder text
        self.placeholder = self.scene.addSimpleText("No results to display")
        self.placeholder.setBrush(Qt.lightGray)
        self.placeholder.setPos(0, 0)
        
        self.setTransformationAnchor(QGraphicsView.AnchorUnderMouse)
        self.setResizeAnchor(QGraphicsView.AnchorUnderMouse)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.setDragMode(QGraphicsView.NoDrag)
        self.setMinimumSize(640, 360)
        self._zoom = 0

    def set_photo(self, pixmap=None):
        if pixmap and not pixmap.isNull():
            self.placeholder.setVisible(False)
            self.setDragMode(QGraphicsView.ScrollHandDrag)
            self.photo_item.setPixmap(pixmap)
            self.setSceneRect(self.photo_item.boundingRect())
            self.fitInView(self.photo_item, Qt.KeepAspectRatio)
            self._zoom = 0
        else:
            self.placeholder.setVisible(True)
            self.setDragMode(QGraphicsView.NoDrag)
            self.photo_item.setPixmap(QPixmap())
            self.setSceneRect(self.placeholder.boundingRect())

    def wheelEvent(self, event):
        if not self.photo_item.pixmap().isNull():
            if event.angleDelta().y() > 0:
                factor = 1.25
                self._zoom += 1
            else:
                factor = 0.8
                self._zoom -= 1
            
            if self._zoom > 0:
                self.scale(factor, factor)
            elif self._zoom == 0:
                self.fitInView(self.photo_item, Qt.KeepAspectRatio)
            else:
                self._zoom = 0

    def resizeEvent(self, event):
        if self._zoom == 0 and not self.photo_item.pixmap().isNull():
            self.fitInView(self.photo_item, Qt.KeepAspectRatio)
        super().resizeEvent(event)


class ResultsTab(QWidget):
    """Results tab UI"""
    
    def __init__(self, config):
        super().__init__()
        self.config = config
        self.current_image = None
        self.adjusted_image = None
        self.current_analysis = None
        self.analysis_data = None
        self.init_ui()
    
    def init_ui(self):
        """Initialize user interface"""
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(20)
        
        # Splitter for resizable sections
        splitter = QSplitter(Qt.Vertical)
        
        # --- Image Preview Section ---
        image_group = QGroupBox("Image Preview")
        image_layout = QVBoxLayout(image_group)
        
        self.image_preview = PhotoViewer()
        image_layout.addWidget(self.image_preview)
        
        # Brightness and Contrast controls
        adj_layout = QHBoxLayout()
        
        # Brightness
        bright_label = QLabel("Brightness:")
        self.brightness_slider = QSlider(Qt.Horizontal)
        self.brightness_slider.setRange(-100, 100)
        self.brightness_slider.setValue(0)
        self.brightness_slider.valueChanged.connect(self.apply_brightness_contrast)
        self.brightness_slider.setEnabled(False)
        adj_layout.addWidget(bright_label)
        adj_layout.addWidget(self.brightness_slider)

        # Contrast
        contrast_label = QLabel("Contrast:")
        self.contrast_slider = QSlider(Qt.Horizontal)
        self.contrast_slider.setRange(-100, 100)
        self.contrast_slider.setValue(0)
        self.contrast_slider.valueChanged.connect(self.apply_brightness_contrast)
        self.contrast_slider.setEnabled(False)
        adj_layout.addWidget(contrast_label)
        adj_layout.addWidget(self.contrast_slider)

        image_layout.addLayout(adj_layout)
        
        splitter.addWidget(image_group)
        
        # --- Details Section ---
        details_group = QGroupBox("Analysis Details")
        details_layout = QHBoxLayout(details_group)
        details_layout.setSpacing(15)

        # Details Text
        self.details_text = QTextEdit()
        self.details_text.setReadOnly(True)
        details_layout.addWidget(self.details_text, 1) # Give more space to text

        # Visualization
        self.figure = Figure(facecolor='#2c3e50')
        self.canvas = FigureCanvas(self.figure)
        details_layout.addWidget(self.canvas, 2) # And to the plot

        splitter.addWidget(details_group)
        main_layout.addWidget(splitter)
        
        # --- Buttons ---
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(15)
        
        self.save_image_button = QPushButton("Save Image")
        self.save_image_button.clicked.connect(self.save_image)
        self.save_image_button.setEnabled(False)
        buttons_layout.addWidget(self.save_image_button)
        
        self.save_data_button = QPushButton("Save Analysis Data")
        self.save_data_button.clicked.connect(self.save_data)
        self.save_data_button.setEnabled(False)
        buttons_layout.addWidget(self.save_data_button)
        
        self.export_button = QPushButton("Export to Excel")
        self.export_button.clicked.connect(self.export_to_excel)
        self.export_button.setEnabled(False)
        buttons_layout.addWidget(self.export_button)
        
        self.analyze_adjusted_button = QPushButton("Analyze Adjusted Image")
        self.analyze_adjusted_button.clicked.connect(self.analyze_adjusted_image)
        self.analyze_adjusted_button.setEnabled(False)
        buttons_layout.addWidget(self.analyze_adjusted_button)
        
        self.clear_button = QPushButton("Clear Results")
        self.clear_button.clicked.connect(self.clear)
        buttons_layout.addWidget(self.clear_button)
        
        main_layout.addLayout(buttons_layout)
    
    @pyqtSlot(object)
    def display_image(self, image):
        """Display an image in the preview"""
        if image is None:
            return
        
        # Store the image
        self.current_image = image.copy()
        
        # Convert to proper format for display
        if len(image.shape) == 3 and image.shape[2] == 3:
            # RGB image
            height, width, channel = image.shape
            bytes_per_line = channel * width
            q_image = QImage(image.data, width, height, bytes_per_line, QImage.Format_RGB888)
        else:
            # Try to handle grayscale image
            height, width = image.shape[:2]
            q_image = QImage(image.data, width, height, width, QImage.Format_Grayscale8)
        
        # Display the QImage in the QLabel
        self.image_preview.set_photo(QPixmap.fromImage(q_image))
        
        # Enable save buttons
        self.save_image_button.setEnabled(True)
        self.analyze_adjusted_button.setEnabled(True)
    
    @pyqtSlot(dict)
    def display_analysis(self, analysis_data):
        """Display analysis results"""
        if not analysis_data:
            return
        
        # Store the analysis data
        self.current_analysis = analysis_data
        
        # Display the brightest frame with highlight
        if 'brightest_frame_marked' in analysis_data:
            marked_frame = analysis_data['brightest_frame_marked']
        else:
            marked_frame = analysis_data['brightest_frame']
        
        # Display the image
        self.current_image = marked_frame.copy()
        self.update_image_preview(self.current_image)
        
        # Update details text
        self.update_details_text(analysis_data)
        
        # Update visualization
        self.update_visualization(analysis_data)
        
        # Enable save buttons
        self.save_image_button.setEnabled(True)
        self.save_data_button.setEnabled(True)
        self.analyze_adjusted_button.setEnabled(True)

        # Reset and enable sliders
        self.brightness_slider.setEnabled(True)
        self.brightness_slider.setValue(0)
        self.contrast_slider.setEnabled(True)
        self.contrast_slider.setValue(0)
    
    def update_details_text(self, analysis_data):
        """Update the details text with analysis information"""
        self.analysis_data = analysis_data  # Store the data for export
        self.export_button.setEnabled(True)  # Enable export button
        
        details = f"""
        <b>Analysis Results</b>
        <hr>
        <b>Timestamp:</b> {analysis_data.get('timestamp', 'N/A')}
        <br>
        <b>Media Path:</b> {analysis_data.get('media_path', 'N/A')}
        <br>
        <br>
        <b>Brightest Point (X, Y):</b> {analysis_data.get('brightest_point', 'N/A')}
        <br>
        <b>Max Brightness (0-255):</b> {analysis_data.get('max_brightness', 'N/A'):.2f}
        <br>
        <b>Average Brightness:</b> {analysis_data.get('average_brightness', 'N/A'):.2f}
        """
        
        if 'frame_number' in analysis_data:
            details += f"<br><b>Frame Number:</b> {analysis_data.get('frame_number', 'N/A')}"
        
        # Add adjustment information if available
        if 'adjustment_metadata' in analysis_data:
            adj_meta = analysis_data['adjustment_metadata']
            if adj_meta.get('was_adjusted', False):
                details += f"""
                <br><br><b>Image Adjustments:</b>
                <br><b>Brightness:</b> {adj_meta.get('brightness_adjustment', 0)}
                <br><b>Contrast:</b> {adj_meta.get('contrast_adjustment', 0)}
                """
            
        self.details_text.setHtml(details)
    
    def update_visualization(self, analysis_data):
        """Update the visualization with brightness histogram"""
        self.figure.clear()
        
        if 'histogram' in analysis_data and analysis_data['histogram'] is not None:
            hist = analysis_data['histogram']
            
            ax = self.figure.add_subplot(111)
            ax.plot(hist, color='#3498db')
            ax.set_title("Brightness Histogram", color='white')
            ax.set_xlabel("Brightness Level", color='white')
            ax.set_ylabel("Pixel Count", color='white')
            ax.grid(True, linestyle='--', alpha=0.6)
            ax.set_facecolor('#34495e')

            # Change tick colors
            ax.tick_params(axis='x', colors='white')
            ax.tick_params(axis='y', colors='white')

            # Change spine colors
            for spine in ax.spines.values():
                spine.set_edgecolor('white')

            self.canvas.draw()
    
    def apply_brightness_contrast(self):
        """Apply brightness and contrast to the current image."""
        if self.current_image is None:
            return

        brightness = self.brightness_slider.value()
        contrast = self.contrast_slider.value()

        # Contrast is a multiplier (alpha), brightness is an offset (beta)
        # Convert slider range to appropriate alpha and beta values
        alpha = 1.0 + (contrast / 100.0)
        beta = brightness

        # Apply the transformation
        self.adjusted_image = cv2.convertScaleAbs(self.current_image, alpha=alpha, beta=beta)
        
        # Update the preview with the adjusted image
        self.update_image_preview(self.adjusted_image)
        
        # Enable analyze button when adjustments are made
        self.analyze_adjusted_button.setEnabled(True)

    def update_image_preview(self, image_to_display):
        """Helper to update the image preview label."""
        if image_to_display is None:
            self.image_preview.set_photo(None)
            return

        if len(image_to_display.shape) == 3 and image_to_display.shape[2] == 3:
            h, w, ch = image_to_display.shape
            q_img = QImage(image_to_display.data, w, h, ch * w, QImage.Format_RGB888)
        else:
            h, w = image_to_display.shape
            q_img = QImage(image_to_display.data, w, h, w, QImage.Format_Grayscale8)
        
        self.image_preview.set_photo(QPixmap.fromImage(q_img))

    @pyqtSlot()
    def save_image(self):
        """Save the current image to a file, letting the user select the folder and filename"""
        image_to_save = self.adjusted_image if self.adjusted_image is not None else self.current_image
        if image_to_save is None:
            QMessageBox.warning(self, "Save Error", "No image to save")
            return
        
        # Get current datetime for filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"brightness_analysis_{timestamp}.{self.config['output']['image_format']}"
        
        # Ask user to select a folder
        folder = QFileDialog.getExistingDirectory(self, "Select Folder to Save Image")
        if not folder:
            return
        
        # Ask user for filename
        filename, ok = QInputDialog.getText(self, "Save Image As", "Enter filename:", text=default_name)
        if not ok or not filename:
            return
        
        # Ensure correct extension
        if not filename.lower().endswith(f".{self.config['output']['image_format']}"):
            filename += f".{self.config['output']['image_format']}"
        file_path = os.path.join(folder, filename)
        
        try:
            # Save the image
            if len(image_to_save.shape) == 3:
                # RGB image
                cv2.imwrite(file_path, cv2.cvtColor(image_to_save, cv2.COLOR_RGB2BGR))
            else:
                # Grayscale image
                cv2.imwrite(file_path, image_to_save)
            QMessageBox.information(self, "Save Complete", f"Image saved to:\n{file_path}")
        except Exception as e:
            QMessageBox.warning(self, "Save Error", f"Error saving image: {e}")

    @pyqtSlot()
    def save_data(self):
        """Save the analysis data to a JSON file, letting the user select the folder and filename"""
        if self.current_analysis is None:
            QMessageBox.warning(self, "Save Error", "No analysis data to save")
            return
        
        # Get current datetime for filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"brightness_analysis_{timestamp}.json"
        
        # Ask user to select a folder
        folder = QFileDialog.getExistingDirectory(self, "Select Folder to Save Analysis Data")
        if not folder:
            return
        
        # Ask user for filename
        filename, ok = QInputDialog.getText(self, "Save Analysis Data As", "Enter filename:", text=default_name)
        if not ok or not filename:
            return
        
        # Ensure correct extension
        if not filename.lower().endswith(".json"):
            filename += ".json"
        file_path = os.path.join(folder, filename)
        
        try:
            # Create a copy of the data, removing non-serializable items
            data_to_save = self.current_analysis.copy()
            # Remove numpy arrays and other non-serializable objects
            keys_to_remove = []
            for key, value in data_to_save.items():
                if isinstance(value, np.ndarray):
                    keys_to_remove.append(key)
            for key in keys_to_remove:
                data_to_save.pop(key)
            # Convert tuple coordinates to lists for JSON serialization
            if 'brightest_point' in data_to_save:
                data_to_save['brightest_point'] = list(data_to_save['brightest_point'])
            # Add timestamp and metadata
            data_to_save['timestamp'] = timestamp
            data_to_save['app_version'] = "1.0.0"
            # Save to JSON file
            with open(file_path, 'w') as f:
                json.dump(data_to_save, f, indent=4)
            QMessageBox.information(self, "Save Complete", f"Analysis data saved to:\n{file_path}")
        except Exception as e:
            QMessageBox.warning(self, "Save Error", f"Error saving data: {e}")
    
    @pyqtSlot()
    def save_results(self):
        """Save both image and data"""
        self.save_image()
        self.save_data()
    
    @pyqtSlot()
    def analyze_adjusted_image(self):
        """Analyze the adjusted image (with brightness/contrast changes)"""
        # Get the image to analyze (adjusted if available, otherwise original)
        image_to_analyze = self.adjusted_image if self.adjusted_image is not None else self.current_image
        
        if image_to_analyze is None:
            QMessageBox.warning(self, "Analysis Error", "No image to analyze")
            return
        
        try:
            # Import the analyzer here to avoid circular imports
            from analysis.brightness_analyzer import BrightnessAnalyzer
            
            # Create analyzer instance
            analyzer = BrightnessAnalyzer(self.config)
            
            # Run analysis on the adjusted image
            analysis_result = analyzer.analyze_image(image_to_analyze)
            
            # Add metadata about the adjustment
            if self.adjusted_image is not None:
                brightness_value = self.brightness_slider.value()
                contrast_value = self.contrast_slider.value()
                analysis_result['adjustment_metadata'] = {
                    'brightness_adjustment': brightness_value,
                    'contrast_adjustment': contrast_value,
                    'was_adjusted': True
                }
            else:
                analysis_result['adjustment_metadata'] = {
                    'brightness_adjustment': 0,
                    'contrast_adjustment': 0,
                    'was_adjusted': False
                }
            
            # Display the new analysis results
            self.display_analysis(analysis_result)
            
            QMessageBox.information(self, "Analysis Complete", 
                                  "Analysis of adjusted image completed successfully!")
            
        except Exception as e:
            QMessageBox.critical(self, "Analysis Error", 
                               f"Error analyzing adjusted image: {str(e)}")
    
    @pyqtSlot()
    def clear(self):
        """Clear the current results"""
        self.current_image = None
        self.adjusted_image = None
        self.current_analysis = None
        self.image_preview.set_photo(None)
        self.details_text.clear()
        self.figure.clear()
        self.canvas.draw()
        
        # Disable buttons
        self.save_image_button.setEnabled(False)
        self.save_data_button.setEnabled(False)
        self.export_button.setEnabled(False)
        self.analyze_adjusted_button.setEnabled(False)

        # Disable and reset sliders
        self.brightness_slider.setEnabled(False)
        self.brightness_slider.setValue(0)
        self.contrast_slider.setEnabled(False)
        self.contrast_slider.setValue(0)
    
    def export_to_excel(self):
        """Export analysis data to Excel"""
        if not self.analysis_data:
            return
            
        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(
                self,
                "Export Error",
                "Excel export requires openpyxl package. Please install it with: pip install openpyxl"
            )
            return
            
        try:
            # Create a new workbook
            wb = Workbook()
            
            # Get the active sheet and rename it
            ws = wb.active
            ws.title = "Brightness Analysis"
            
            # Prepare analysis data
            data = [
                ('Maximum Brightness', self.analysis_data.get('max_brightness', 0)),
                ('Average Brightness', self.analysis_data.get('average_brightness', 0)),
                ('Brightest Point X', self.analysis_data.get('brightest_point', (0, 0))[0]),
                ('Brightest Point Y', self.analysis_data.get('brightest_point', (0, 0))[1]),
                ('Frame Number', self.analysis_data.get('frame_number', 0)),
                ('Analysis Time', self.analysis_data.get('metadata', {}).get('analysis_time', ''))
            ]
            
            # Add video-specific data if available
            if self.analysis_data.get('is_video', False):
                data.extend([
                    ('Total Frames', self.analysis_data.get('total_frames', 0)),
                    ('FPS', self.analysis_data.get('fps', 0)),
                    ('Frame Time (seconds)', self.analysis_data.get('frame_number', 0) / self.analysis_data.get('fps', 1))
                ])
            
            # Write headers horizontally
            for col, (metric, _) in enumerate(data, 1):
                cell = ws.cell(row=1, column=col, value=metric)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            # Write values horizontally
            for col, (_, value) in enumerate(data, 1):
                ws.cell(row=2, column=col, value=value)
            
            # Add histogram data to a new sheet if available
            if 'histogram' in self.analysis_data and self.analysis_data['histogram'] is not None:
                hist_ws = wb.create_sheet("Histogram")
                
                # Add headers
                hist_headers = ['Brightness Level', 'Pixel Count']
                for col, header in enumerate(hist_headers, 1):
                    cell = hist_ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # Add histogram data
                histogram = self.analysis_data['histogram']
                for i, count in enumerate(histogram):
                    hist_ws.cell(row=i+2, column=1, value=i)
                    hist_ws.cell(row=i+2, column=2, value=int(count))
            
            # Get save file path
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            default_name = f'brightness_analysis_{timestamp}.xlsx'
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Save Excel File",
                default_name,
                "Excel Files (*.xlsx)"
            )
            
            if file_path:
                # Save the workbook
                wb.save(file_path)
                
                QMessageBox.information(
                    self,
                    "Export Successful",
                    f"Data has been exported to:\n{file_path}"
                )
                
        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Error",
                f"Failed to export data: {str(e)}"
            )
